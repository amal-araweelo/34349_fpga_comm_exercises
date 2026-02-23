# This script was generated using chatgpt


import openpyxl

# ---- CONFIG ----
FILE_PATH = "8_bit_parallel_state_matrix.xlsx"
SHEET_NAME = "Sheet1"

HEADER_ROW = 2
ROW_LABEL_COLUMN = 1
DATA_START_ROW = 3
DATA_START_COL = 2
# -----------------


def format_signal_sv(name):
    """
    R24 -> crc_r[24]
    M1  -> data_in[1]
    """
    if name.startswith("R"):
        return f"crc_r[{name[1:]}]"
    elif name.startswith("M"):
        return f"data_in[{name[1:]}]"
    else:
        return name


wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
ws = wb[SHEET_NAME]

# Read column headers
col_headers = []
col = DATA_START_COL
while ws.cell(row=HEADER_ROW, column=col).value:
    col_headers.append(ws.cell(row=HEADER_ROW, column=col).value)
    col += 1

# Read row labels
row_labels = []
row = DATA_START_ROW
while ws.cell(row=row, column=ROW_LABEL_COLUMN).value:
    row_labels.append(ws.cell(row=row, column=ROW_LABEL_COLUMN).value)
    row += 1
    
print("// -----------------------------")
print("// Auto-generated CRC equations")
print("// -----------------------------\n")

for col_index, col_name in enumerate(col_headers):
    xor_terms = []

    for row_index, row_name in enumerate(row_labels):
        cell_value = ws.cell(
            row=DATA_START_ROW + row_index,
            column=DATA_START_COL + col_index
        ).value

        if cell_value == 1:
            xor_terms.append(format_signal_sv(row_name))

    if xor_terms:
        left_side = format_signal_sv(col_name)
        right_side = " ^ ".join(xor_terms)
        print(f"{left_side} <= {right_side};")
    else:
        # If no XOR terms → assign zero
        left_side = format_signal_sv(col_name)
        print(f"{left_side} <= '0;")